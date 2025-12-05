from datetime import datetime

from openpyxl import Workbook

def sheet(classes: list, start_date: str, end_date: str, file_location: str) -> None:
    """
    Generates the hour sheet file and saves it to a given file location.

    :param list classes: List with the names of the classes the file should be generated with.
    :param str start_date: String representation of thed ate the sheet should start at. Using '-' as a seperator.
    :param str end_date: String reperesentation of the date the sheet should end at. Using '-' as a seperator.
    :param str file_location: Location the file should be stored at.
    """

    start_date = datetime.strptime(start_date, "%d-%m-%Y")
    end_date = datetime.strptime(end_date, "%d-%m-%Y")

    if end_date < start_date:
        raise ValueError("Start date is later than end date.")
    
    days = (end_date - start_date).days + 1

    wb = Workbook()
    ws = wb.active
    ws.title = "Hoursheet"

    add_header(ws)

    wb.save(file_location)


def add_header(sheet, row_offset: int = 2, col_offset: int = 3, sessions: int = 6):
    """
    Adds the header to the sheet.

    :param sheet: Sheet to which the header should be added.
    :param int row_offset: Row offset where the header should start relative to the top left of the spreadsheet.
    :param int col_offset: Column offset where the header should start relative to the top left of the spreadsheet.
    :param int sessions: Amount of sessions blocks should be generated.
    """

    for i in range(sessions - 1):
        sheet.cell(row=row_offset, column=col_offset + i*3).value = "Start"
        sheet.cell(row=row_offset, column=col_offset + i*3 + 1).value = "Stop"
        sheet.cell(row=row_offset, column=col_offset + i*3 + 2).value = "Course"

    


def format_date(date: str) -> str:
    """
    Ensures the time is formatted according to the format specifiers of the Python time library.
    Assumes the year is in the 2000 century if the year is two characters long.
    Assumes the year is between 0-9999.

    :param str date: A string representation of a date using '-' as delimiter. 
    """

    if date.count("-") != 2:
        raise ValueError("Please provide the date in a valid string format (with - as seperator)")
    
    res = ""
    for i, p in enumerate(date.split("-")):
        if i == 2:
            if len(p) == 2:
                return res + "20" + p
            else:
                return res + (4 - len(p))*"0" + p
            
        res += (2 - len(p))*"0" + p + "-"
        



def test_format_date():
    print(20*"=" + " Testing format_date() " + 20*"=" + "\n")

    dates = [
        "28-2-2025",
        "6-7-13",
        "13-8-813",
        "6-6-6",
        "25-12-2000"
    ]

    for d in dates:
        print(f"{d} becomes {format_date(d)}")

    print("\n" + 63*"=" + "\n\n")

def test_sheet():
    sheet(["Computer architecture", "Operating systems", "Computer networks"], "1-1-2000", "11-1-2000", "test.xlsx")

if __name__ == "__main__":
    # test_format_date()
    test_sheet()